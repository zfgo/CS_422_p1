# created 2023-04-22
# updated 2023-04-23 : add to_json fxn
# updated 2023-04-29 : add id to the Document model (used to get the
#   correct file during download)
from django.db import models
import csv
import json
import openpyxl
from .validators import validate_file_extension

class Task(models.Model):
    task_desc = models.CharField(max_length=255, blank=True, null=True)
    period = models.FloatField(null=True)
    n_forecasts = models.IntegerField(null=True)
    id = models.BigAutoField(primary_key=True)

def document_upload_path(instance, filename):
    print("")
    return f"documents/{instance.task.id}/{filename}"

def metrics_upload_path(instance, filename):
    print(filename)
    return f"documents/{instance.task.id}/metrics/{filename}"

class Document(models.Model):
    """Document model that stores all data associated with a file. this includes the json data and weather
    its test training or contributor. """
    task = models.ForeignKey(Task, on_delete=models.CASCADE, null=True) # files are associated with a TS forecasting task
    description = models.CharField(max_length=255, blank=True)
    if_test = models.BooleanField(default=True, null=True) # True for test data, False for train data
    document = models.FileField(upload_to=document_upload_path, validators=[validate_file_extension], null=True)
    document2 = models.FileField(upload_to=None, validators=[validate_file_extension], null=True)
    uploaded_at = models.DateTimeField(auto_now_add=True)
    json_data = models.JSONField(null=True)
    id = models.BigAutoField(primary_key=True)
    file_name = models.CharField(max_length=255, blank=True, null=True) #the file name 
    fname = models.CharField(max_length=255, blank=True, null=True)
    fdescription = models.CharField(max_length=255, blank=True, null=True)
    funits = models.CharField(max_length=255, blank=True, null=True)
    fvector_size = models.IntegerField(null=True)
    flength = models.IntegerField(null=True)
    fsampling_period = models.FloatField(null=True)
    image = models.ImageField(upload_to=metrics_upload_path, null=True)
    solution = models.FileField(upload_to=metrics_upload_path, null=True)


    def to_json(self):
        """converts file to json"""
        path_str_list = self.document.name.split('.')

        if path_str_list[-1].lower() == 'csv':
            self.json_data = csv_to_json(self.document.path)

        elif path_str_list[-1].lower() == 'xlsx':
            self.json_data = xlsx_to_json(self.document.path)

        #print(self.document.path)
        #self.json_data = make_json(self.document.path)


    #def __str__(self):
    #   return self.description

    def json_to_csv(self):
        """
        Method to convert the json data stored into the model to a CSV 
        file to be downloaded by the user
        """
        # create the file name
        csv_file_name = "downloads/" + self.document.name
        
        # Load the JSON data and extract the header row
        data = json.loads(self.json_data)
        header = list(data[0].keys())

        # Open the CSV file for writing and write the header row
        with open(csv_file_name, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(header)

            # Write each row of data to the CSV file
            for row in data:
                writer.writerow([row[column] for column in header])

    def json_to_xlsx(self):
        # create the file name
        xlsx_file_name = "downloads/" + self.document.name
        # Load the JSON data and extract the header row
        data = json.loads(self.json_data)
        header = list(data[0].keys())

        # Create a new Excel workbook and worksheet
        wb = openpyxl.Workbook()
        sheet_name = "Sheet1"
        sheet = wb.active
        sheet.title = sheet_name

        # Write the header row to the worksheet
        sheet.append(header)

        # Write each row of data to the worksheet
        for row in data:
            sheet.append([row[column] for column in header])

        # Save the Excel workbook to a file
        wb.save(xlsx_file_name)

    def __str__(self):
        return self.description


def csv_to_json(csv_file_name, time_column_index=1):
    """
    Function to convert a CSV to JSON

    Take the path to the csv file and the column that holds time, and 
    convert to a JSON object. We assume that the time column is 1-based
    """
    time_column_index -= 1 # it is entered as 1-based, so we convert to 0-based indexing

    # Open the CSV file and read its contents
    with open(csv_file_name, 'rt') as csv_file:
        reader = csv.reader(csv_file)

        # Create a list to store the rows as dictionaries
        data = []
        header = []

        # Iterate through the CSV file and convert each row to a dictionary
        for row in reader:
            if not header:
                header = row
            else:
                row_data = {}
                for i in range(len(header)):
                    if i == time_column_index:
                        row_data[header[i]] = float(row[i])
                    else:
                        row_data[header[i]] = float(row[i])
                data.append(row_data)

    # Convert the list of dictionaries to a JSON object and return it
    return json.dumps(data)


def xlsx_to_json(xlsx_file_name):
    """
    Function to convert a XLSX (Excel) to JSON

    Take the path to the xlsx file and convert to a JSON object. We use 
    the openpyxl library to help us do this.
    """
    # Open the Excel file and select the first sheet
    wb = openpyxl.load_workbook(xlsx_file_name)
    sheet = wb.active

    # Create a list to store the rows as dictionaries
    data = []
    header = []

    # Iterate through the sheet and convert each row to a dictionary
    for row in sheet.iter_rows(values_only=True):
        if not header:
            header = row
        else:
            row_data = {}
            for i in range(len(header)):
                row_data[header[i]] = row[i]
            data.append(row_data)

    # Convert the list of dictionaries to a JSON object and return it
    return json.dumps(data)
